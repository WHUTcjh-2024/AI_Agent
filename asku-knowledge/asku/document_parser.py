"""Offline, bounded format-aware extraction. Binary decoding is never a fallback."""

from __future__ import annotations

import hashlib
import io
import os
import re
import subprocess
import tempfile
import time
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path

from .admission import canonical_text

PARSER_VERSION = "local-parser-v1"
MAX_FILE_BYTES = 64 * 1024 * 1024
MAX_ARCHIVE_BYTES = 128 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 100
MAX_TEXT_CHARS = 1_000_000


@dataclass
class Parsed:
    status: str
    format: str
    text: str = ""
    parser: str = PARSER_VERSION
    errors: list[str] = field(default_factory=list)
    pages: int = 0
    ocr_pages: int = 0
    members: list[dict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return asdict(self)


def finish(result: Parsed) -> Parsed:
    result.text = result.text.replace("\x00", "").strip()
    compact = canonical_text(result.text)
    if not compact:
        result.status = "EMPTY" if not result.errors else "FAILED"
    elif len(result.text) > MAX_TEXT_CHARS:
        result.text = result.text[:MAX_TEXT_CHARS]
        result.status = "REVIEW"
        result.errors.append("text_budget_exceeded")
    elif len(compact) < 100:
        result.status = "REVIEW"
        result.errors.append("short_text")
    if "\ufffd" in result.text:
        result.status = "REVIEW"
        result.errors.append("replacement_characters")
    return result


def markdown_table(rows) -> str:
    rows = [
        [
            str(v if v is not None else "")
            .replace("|", "\\|")
            .replace("\n", "<br>")
            .strip()
            for v in row
        ]
        for row in rows
    ]
    rows = [r for r in rows if any(r)]
    if not rows:
        return ""
    width = max(map(len, rows))
    rows = [r + [""] * (width - len(r)) for r in rows]
    return "\n".join(
        ["| " + " | ".join(rows[0]) + " |", "| " + " | ".join(["---"] * width) + " |"]
        + ["| " + " | ".join(r) + " |" for r in rows[1:]]
    )


def sniff(data: bytes) -> str:
    head = data[:2048].lstrip()
    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(b"PK"):
        return "zip"
    if head.startswith(b"Rar!"):
        return "rar"
    if head.startswith(bytes.fromhex("d0cf11e0a1b11ae1")):
        return "ole"
    if head.startswith(b"{\\rtf"):
        return "rtf"
    if re.match(rb"(?is)(<!doctype\s+html|<html|<head|<body|<script)", head):
        return "html"
    if head.startswith((b"\x89PNG", b"\xff\xd8\xff", b"GIF8", b"II*\x00", b"MM\x00*")):
        return "image"
    return "unknown"


def ocr_image(image: bytes, tesseract: str, *, timeout: int = 60) -> str:
    with tempfile.TemporaryDirectory(prefix="asku-ocr-") as temporary:
        path = Path(temporary) / "page.png"
        path.write_bytes(image)
        result = subprocess.run(
            [tesseract, str(path), "stdout", "-l", "chi_sim+eng", "--psm", "3"],
            capture_output=True,
            timeout=timeout,
            check=True,
            env={**os.environ, "OMP_THREAD_LIMIT": "1"},
        )
        return result.stdout.decode("utf-8", errors="strict")


def parse_pdf(data: bytes, tesseract: str | None) -> Parsed:
    import pymupdf

    result = Parsed("PARSED", "pdf")
    parts = []
    started = time.monotonic()
    with pymupdf.open(stream=data, filetype="pdf") as document:
        if document.needs_pass:
            return Parsed("FAILED", "pdf", errors=["encrypted_pdf"])
        result.pages = len(document)
        for index, page in enumerate(document):
            if time.monotonic() - started > 180:
                result.errors.append("document_time_budget_exceeded")
                result.status = "REVIEW"
                break
            if index >= 300:
                result.errors.append("page_budget_exceeded")
                result.status = "REVIEW"
                break
            text = page.get_text("text", sort=True)
            # Images with little embedded text must not count as parsed pages.
            scanned = len(canonical_text(text)) < 30 and bool(page.get_images())
            if scanned:
                if not tesseract:
                    result.errors.append(f"page_{index + 1}:ocr_required")
                    result.status = "REVIEW"
                else:
                    pix = page.get_pixmap(dpi=160, alpha=False)
                    if pix.width * pix.height > 25_000_000:
                        result.errors.append(f"page_{index + 1}:image_budget_exceeded")
                        result.status = "REVIEW"
                    else:
                        text = ocr_image(pix.tobytes("png"), tesseract)
                        result.ocr_pages += 1
                        # OCR recovers evidence but does not certify characters or table alignment.
                        result.status = "REVIEW"
            elif text.strip():
                tables = page.find_tables(strategy="lines").tables
                if tables:
                    outside = [
                        b[4]
                        for b in page.get_text("blocks", sort=True)
                        if len(b) > 4
                        and isinstance(b[4], str)
                        and not any(
                            pymupdf.Rect(b[:4]).intersects(pymupdf.Rect(t.bbox))
                            for t in tables
                        )
                    ]
                    text = "\n\n".join(
                        outside + [markdown_table(t.extract()) for t in tables]
                    )
            if text.strip():
                parts.append(f"## 第 {index + 1} 页\n\n{text.strip()}")
        result.text = "\n\n".join(parts)
    if result.ocr_pages:
        result.errors.append("ocr_requires_review")
    return finish(result)


def parse_bytes(
    data: bytes,
    filename: str,
    *,
    tesseract: str | None = None,
    legacy_dir: Path | None = None,
    archive_dir: Path | None = None,
    depth: int = 0,
) -> Parsed:
    kind = sniff(data)
    if len(data) > MAX_FILE_BYTES:
        return Parsed("FAILED", kind, errors=["file_budget_exceeded"])
    if not data:
        return Parsed("EMPTY", kind)
    try:
        if kind == "rar":
            cached = (
                archive_dir / (hashlib.sha256(data).hexdigest() + ".zip")
                if archive_dir
                else None
            )
            if not cached or not cached.is_file():
                return Parsed("FAILED", "rar", errors=["rar_conversion_required"])
            result = parse_bytes(
                cached.read_bytes(),
                filename + ".zip",
                tesseract=tesseract,
                legacy_dir=legacy_dir,
                archive_dir=archive_dir,
                depth=depth,
            )
            result.format = "rar"
            return result
        if kind == "pdf":
            return parse_pdf(data, tesseract)
        if kind == "image":
            if not tesseract:
                return Parsed("FAILED", kind, errors=["ocr_required"])
            from PIL import Image

            with Image.open(io.BytesIO(data)) as im:
                if im.width * im.height > 25_000_000:
                    return Parsed("FAILED", kind, errors=["image_budget_exceeded"])
                buffer = io.BytesIO()
                im.convert("RGB").save(buffer, format="PNG")
            return finish(
                Parsed(
                    "REVIEW",
                    kind,
                    ocr_image(buffer.getvalue(), tesseract),
                    errors=["ocr_requires_review"],
                    ocr_pages=1,
                )
            )
        if kind == "zip":
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                members = archive.infolist()
                if (
                    len(members) > MAX_ARCHIVE_MEMBERS * 10
                    or sum(m.file_size for m in members) > MAX_ARCHIVE_BYTES
                ):
                    return Parsed("FAILED", kind, errors=["archive_budget_exceeded"])
                names = set(archive.namelist())
                if "word/document.xml" in names:
                    from docx import Document

                    doc = Document(io.BytesIO(data))
                    from docx.table import Table
                    from docx.text.paragraph import Paragraph

                    parts = [
                        block.text
                        if isinstance(block, Paragraph)
                        else markdown_table(
                            [[c.text for c in r.cells] for r in block.rows]
                        )
                        for block in doc.iter_inner_content()
                        if isinstance(block, (Table, Paragraph))
                    ]
                    return finish(Parsed("PARSED", "docx", "\n\n".join(parts)))
                if "xl/workbook.xml" in names:
                    import openpyxl

                    workbook = openpyxl.load_workbook(
                        io.BytesIO(data),
                        read_only=True,
                        data_only=False,
                        keep_links=False,
                    )
                    try:
                        parts = []
                        for sheet in workbook:
                            if sheet.max_row * sheet.max_column > 500_000:
                                return Parsed(
                                    "FAILED", "xlsx", errors=["cell_budget_exceeded"]
                                )
                            parts.append(
                                "## "
                                + sheet.title
                                + "\n\n"
                                + markdown_table(sheet.iter_rows(values_only=True))
                            )
                        return finish(Parsed("PARSED", "xlsx", "\n\n".join(parts)))
                    finally:
                        workbook.close()
                if depth >= 2 or len(members) > MAX_ARCHIVE_MEMBERS:
                    return Parsed(
                        "FAILED",
                        "archive",
                        errors=["archive_depth_or_members_exceeded"],
                    )
                result = Parsed("PARSED", "archive")
                texts = []
                for member in members:
                    if sum(map(len, texts)) > MAX_TEXT_CHARS:
                        result.errors.append("archive_text_budget_exceeded")
                        result.status = "REVIEW"
                        break
                    if member.is_dir() or member.filename.startswith("__MACOSX/"):
                        continue
                    if member.flag_bits & 1 or member.file_size > MAX_FILE_BYTES:
                        child = Parsed(
                            "FAILED",
                            "unknown",
                            errors=["encrypted_or_oversized_member"],
                        )
                    else:
                        child = parse_bytes(
                            archive.read(member),
                            member.filename,
                            tesseract=tesseract,
                            legacy_dir=legacy_dir,
                            archive_dir=archive_dir,
                            depth=depth + 1,
                        )
                    # Never extract member paths onto the filesystem.
                    result.members.append(
                        {
                            "name": member.filename,
                            "status": child.status,
                            "format": child.format,
                            "errors": child.errors,
                        }
                    )
                    if child.status != "PARSED":
                        result.status = "REVIEW"
                        result.errors.append("archive_member_not_parsed")
                    if child.text:
                        texts.append(
                            "## 附件成员 " + member.filename + "\n\n" + child.text
                        )
                    result.ocr_pages += child.ocr_pages
                result.text = "\n\n".join(texts)
                return finish(result)
        if kind == "ole":
            import olefile

            with olefile.OleFileIO(io.BytesIO(data)) as ole:
                if ole.exists("Workbook") or ole.exists("Book"):
                    import xlrd

                    book = xlrd.open_workbook(file_contents=data)
                    try:
                        parts = []
                        for sheet in book.sheets():
                            if sheet.nrows * sheet.ncols > 500_000:
                                return Parsed(
                                    "FAILED", "xls", errors=["cell_budget_exceeded"]
                                )
                            parts.append(
                                "## "
                                + sheet.name
                                + "\n\n"
                                + markdown_table(
                                    sheet.row_values(i) for i in range(sheet.nrows)
                                )
                            )
                        return finish(Parsed("PARSED", "xls", "\n\n".join(parts)))
                    finally:
                        book.release_resources()
                is_word = ole.exists("WordDocument")
            if is_word and legacy_dir:
                path = legacy_dir / (hashlib.sha256(data).hexdigest() + ".txt")
                if path.is_file():
                    return finish(
                        Parsed(
                            "PARSED",
                            "doc",
                            path.read_text(encoding="utf-8"),
                            parser=PARSER_VERSION + ":antiword",
                        )
                    )
            return Parsed(
                "FAILED",
                "doc" if is_word else kind,
                errors=["legacy_conversion_required"],
            )
        if kind == "rtf":
            from striprtf.striprtf import rtf_to_text

            return finish(Parsed("PARSED", kind, rtf_to_text(data.decode("latin1"))))
        if kind == "html":
            from bs4 import UnicodeDammit

            from .normalizer import normalize_html

            html = UnicodeDammit(data, is_html=True).unicode_markup
            return finish(
                Parsed(
                    "REVIEW",
                    kind,
                    normalize_html(html or "").markdown,
                    errors=["html_attachment_requires_review"],
                )
            )
        if (
            Path(filename).suffix.lower() in {".txt", ".csv", ".md"}
            and b"\x00" not in data
        ):
            return finish(
                Parsed("PARSED", "text", data.decode("utf-8-sig", errors="strict"))
            )
        return Parsed("FAILED", kind, errors=["unsupported_format"])
    except Exception as exc:
        # Persist diagnostic classes, not document text or secrets from parser errors.
        return Parsed("FAILED", kind, errors=[type(exc).__name__])


def parse_file(path: Path, **options) -> Parsed:
    if not path.is_file():
        return Parsed("FAILED", "missing", errors=["raw_file_missing"])
    if path.stat().st_size > MAX_FILE_BYTES:
        return Parsed("FAILED", "unknown", errors=["file_budget_exceeded"])
    return parse_bytes(path.read_bytes(), path.name, **options)
